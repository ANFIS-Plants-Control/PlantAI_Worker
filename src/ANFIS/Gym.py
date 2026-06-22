from openpyxl import load_workbook
from pathlib import Path
from src.ANFIS.ANFIS import ANFIS
from src.ANFIS.GeneticAlgorithm import GeneticAlgorithm

DATASET_PATH = Path(__file__).resolve().parent / "static" / "MarkeredDatas.xlsx"
MODEL_PARAMETERS_PATH = Path(__file__).resolve().parent / "static" / "model.txt"

class Gym:
    def __init__(self, anfis: ANFIS):
        self.anfis = anfis

    def train(self):
        book = load_workbook(DATASET_PATH)
        ws = book.active
        trainingParams, targets = self._load_training_data(ws)
        
        ga = GeneticAlgorithm(100, 10, 0.01, len(self.anfis.getParameters()), self.anfis)

        newParams = ga.runAlgorithm(100, trainingParams, targets, 0.2)

        try:
            with open(MODEL_PARAMETERS_PATH, "x") as file:
                for np in newParams:
                    file.write(str(f'{np}\n'))
                print("Файл успешно создан!")
        except FileExistsError:
            print("Файл уже существует.")

        self.anfis.setParameters(newParams)

        return self.anfis

    @staticmethod
    def _load_training_data(ws) -> tuple[list[dict[str, float]], list[float]]:
        trainingParams: list[dict[str, float]] = []
        targets: list[float] = []
        skip = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if(skip != 10): 
                skip = 0
                continue
            if row[0] is None or row[1] is None or row[2] is None or row[3] is None:
                continue

            co2 = float(row[0])
            temp = float(row[1])
            hum = float(row[2])
            target = float(row[3])

            trainingParams.append({
                "temperature": temp,
                "humidity": hum,
                "co2": co2,
            })
            targets.append(target)

            skip += 1
        return trainingParams, targets
